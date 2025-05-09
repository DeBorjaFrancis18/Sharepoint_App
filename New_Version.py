import os
import requests
import pandas as pd
from datetime import datetime
import fnmatch
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
import sys

def show_popup(title, message):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()

def read_config_file(file_path):
    config_values = {}
    try:
        with open(file_path, 'r') as file:
            for line in file:
                parts = line.strip().split('=')
                if len(parts) >= 2:
                    key = parts[0].strip()
                    value = '='.join(parts[1:]).strip().strip('"')
                    config_values[key] = value
                else:
                    print(f"Skipping malformed line: {line.strip()}")
    except FileNotFoundError:
        error_msg = f"Config file '{file_path}' not found."
        print(error_msg)
        show_popup("Error", error_msg)
        sys.exit(1)
    except Exception as e:
        error_msg = f"Error reading config file: {str(e)}"
        print(error_msg)
        show_popup("Error", error_msg)
        sys.exit(1)
    return config_values

def update_log_sheet(log_sheet, file_name, status):
    log_sheet.insert_rows(2)
    log_sheet['A2'] = file_name
    log_sheet['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet['C2'] = status

def log_result(log_file_path, file_name, status):
    if os.path.exists(log_file_path):
        workbook = load_workbook(log_file_path)
        log_sheet = workbook.active
    else:
        workbook = Workbook()
        log_sheet = workbook.active
        log_sheet.append(['File Name', 'Timestamp', 'Status'])
    
    update_log_sheet(log_sheet, file_name, status)
    workbook.save(log_file_path)

def is_file_large(file_path, max_size_mb=250):
    
    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
    return file_size_mb > max_size_mb

def get_sharepoint_context_using_app(config_values):
    sharepoint_url = config_values.get('DestinationSiteURL')
    client_credentials = ClientCredential(
        config_values.get('Client Id'), 
        config_values.get('Client Secret')
    )
    ctx = ClientContext(sharepoint_url).with_credentials(client_credentials)
    return ctx

def get_config_values(file_path=None):

    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    config_file_path = os.path.join(script_dir, "config.txt")
    
    if not os.path.exists(config_file_path):
        error_msg = f"ERROR: Config file not found at {config_file_path}"
        print(error_msg)
        show_popup("Config File Error", error_msg)
        sys.exit(1)

    config_values = read_config_file(config_file_path)
    
    if file_path:
        config_values['SourceFolderPath'] = os.path.dirname(file_path)
        config_values['FileName'] = os.path.basename(file_path)
    else:
        if not config_values.get('SourceFolderPath') or not config_values.get('FileName'):
            error_msg = "Source folder path or file name pattern is missing in the config file."
            print(error_msg)
            show_popup("Config File Error", error_msg)
            sys.exit(1)

    config_values['TargetFolderURL'] = config_values.get('DestinationFolderURL')
    config_values['LogFilePath'] = config_values.get('LogFilePath')

    return config_values

def upload_small_files(file_path, config_values):
    ctx = get_sharepoint_context_using_app(config_values)
    target_folder_url = config_values['TargetFolderURL']
    target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)

    log_file_path = config_values['LogFilePath']
    log_workbook = load_workbook(log_file_path) if log_file_path and os.path.exists(log_file_path) else None
    log_sheet = log_workbook.active if log_workbook else None

    print(f"Target folder URL: {target_folder_url}")

    try:
        file_name = os.path.basename(file_path)
        print(f"\nProcessing file: {file_name}")
        with open(file_path, 'rb') as file_content:
            target_folder.upload_file(file_name, file_content).execute_query()
            print(f"File '{file_name}' uploaded successfully.")
            if log_sheet:
                update_log_sheet(log_sheet, file_name, "Successful")
                log_workbook.save(log_file_path)
    except Exception as e:
        error_msg = f"Failed to upload {file_name}: {str(e)}"
        print(error_msg)
        if log_sheet:
            update_log_sheet(log_sheet, file_name, "Failed")
            log_workbook.save(log_file_path)

def upload_large_files(file_path, config_values):
    """
    Uploads large files to SharePoint in chunks of 50MB.

    Args:
        file_path (str): Path to the file to upload.
        config_values (dict): Configuration values for SharePoint and logging.
    """
    ctx = get_sharepoint_context_using_app(config_values)
    target_folder_url = config_values['TargetFolderURL']
    target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)

    log_file_path = config_values['LogFilePath']
    log_workbook = load_workbook(log_file_path) if log_file_path and os.path.exists(log_file_path) else None
    log_sheet = log_workbook.active if log_workbook else None

    print(f"Target folder URL: {target_folder_url}")

    try:
        file_name = os.path.basename(file_path)
        print(f"\nProcessing large file: {file_name}")

        # Open the file and upload in chunks
        file_size = os.path.getsize(file_path)
        chunk_size = 50 * 1024 * 1024  # 50MB
        offset = 0  # Initialize offset to track progress

        with open(file_path, 'rb') as file_content:
            # Start the upload session
            first_chunk = file_content.read(chunk_size)
            uploaded_file = target_folder.files.add(file_name, first_chunk, overwrite=True).execute_query()
            upload_id = uploaded_file.unique_id

            offset += len(first_chunk)
            print(f"Uploaded first chunk: {offset}/{file_size} bytes")

            # Continue uploading chunks
            while offset < file_size:
                chunk_data = file_content.read(chunk_size)
                if not chunk_data:
                    break

                if offset + len(chunk_data) < file_size:
                    # Continue uploading intermediate chunks
                    uploaded_file = uploaded_file.continue_upload(upload_id, offset, chunk_data).execute_query()
                else:
                    # Finalize the upload with the last chunk
                    uploaded_file = uploaded_file.finish_upload(upload_id, offset, chunk_data).execute_query()

                offset += len(chunk_data)
                print(f"Uploaded chunk: {offset}/{file_size} bytes")

        print(f"Large file '{file_name}' uploaded successfully.")
        if log_sheet:
            update_log_sheet(log_sheet, file_name, "Successful")
            log_workbook.save(log_file_path)
    except Exception as e:
        error_msg = f"Failed to upload large file '{file_name}': {str(e)}"
        print(error_msg)
        if log_sheet:
            update_log_sheet(log_sheet, file_name, "Failed")
            log_workbook.save(log_file_path)
        raise

if __name__ == "__main__":
    success_count = 0
    failure_count = 0
    processed_files = []

    if len(sys.argv) > 1:
        # Process a single file provided as an argument
        file_path = sys.argv[1]
        config_values = get_config_values(file_path)
        if is_file_large(file_path):
            print(f"The file '{file_path}' is large. Executing 'upload_large_files'.")
            try:
                upload_large_files(file_path, config_values)
                success_count += 1  # Increment only if no exception occurs
                processed_files.append(f"Success (Large): {file_path}")
            except Exception as e:
                failure_count += 1
                processed_files.append(f"Failed (Large): {file_path} - {str(e)}")
        else:
            print(f"The file '{file_path}' is small. Executing 'upload_small_files'.")
            try:
                upload_small_files(file_path, config_values)
                success_count += 1  # Increment only if no exception occurs
                processed_files.append(f"Success: {file_path}")
            except Exception as e:
                failure_count += 1
                processed_files.append(f"Failed: {file_path} - {str(e)}")
    else:
        # Process all files in the source folder
        print("No file path provided. Processing all files in the source folder.")
        config_values = get_config_values()
        source_folder_path = config_values['SourceFolderPath']
        wildcard_pattern = config_values['FileName']
        
        for file_name in os.listdir(source_folder_path):
            if fnmatch.fnmatch(file_name, wildcard_pattern):
                file_path = os.path.join(source_folder_path, file_name)
                if is_file_large(file_path):
                    print(f"The file '{file_name}' is large. Executing 'upload_large_files'.")
                    try:
                        upload_large_files(file_path, config_values)
                        success_count += 1  # Increment only if no exception occurs
                        processed_files.append(f"Success (Large): {file_name}")
                    except Exception as e:
                        failure_count += 1
                        processed_files.append(f"Failed (Large): {file_name} - {str(e)}")
                else:
                    print(f"The file '{file_name}' is small. Executing 'upload_small_files'.")
                    try:
                        upload_small_files(file_path, config_values)
                        success_count += 1  # Increment only if no exception occurs
                        processed_files.append(f"Success: {file_name}")
                    except Exception as e:
                        failure_count += 1
                        processed_files.append(f"Failed: {file_name} - {str(e)}")

    # Show summary popup
    summary_message = (
        f"Execution completed!\n\n"
        f"Files processed: {len(processed_files)}\n"
        f"Success: {success_count}\n"
        f"Failed: {failure_count}\n\n"
        f"Details:\n" + "\n".join(processed_files)
    )
    print(summary_message)  # Print summary to console
    show_popup("Execution Summary", summary_message)
